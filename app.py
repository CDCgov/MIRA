# -*- coding: utf-8 -*-

# Run this app with `python app.py` and
# visit http://127.0.0.1:8050/ in your web browser.


import dash
from dash import dcc, dash_table, html  # , DiskcacheManager   , ctx
import dash_bootstrap_components as dbc


from dash.dependencies import Input, Output, State
import dash_daq as daq

# from plotly.subplots import make_subplots
import plotly.graph_objects as go

# import plotly.express as px
import plotly.io as pio
import pandas as pd

# from math import ceil  # , log10, sqrt
import itertools
from sys import path, argv
from os.path import dirname, realpath  # , isdir, isfile
import os
import yaml
import json
from glob import glob

# from numpy import arange
import base64
import io
import datetime
from openpyxl.worksheet.datavalidation import DataValidation as DV
from openpyxl.worksheet.table import Table, TableStyleInfo
import openpyxl as xl
import re
import requests
import subprocess

path.append(dirname(realpath(__file__)) + "/scripts/")
import conditional_color_range_perCol  # type: ignore

with open(argv[1], "r") as y:
    CONFIG = yaml.safe_load(y)
data_root = CONFIG["DATA_ROOT"]
DEBUG = CONFIG["DEBUG"]
DEPLOY = CONFIG["DEPLOY"]
AVAILABLE_VERSION = CONFIG["VERSION_URL"]
#if DEPLOY:
check_version_interval = 1000 * 60 * 60 * 24  # milliseconds in 1 day
#else:
#    check_version_interval = 3000

app = dash.Dash(external_stylesheets=[dbc.themes.FLATLY])
app.title = "MIRA"
app.config["suppress_callback_exceptions"] = True
app.update_title = None

previousClick = 0


@app.callback(Output("select_run", "options"), Input("run-refresh-button", "n_clicks"))
def refreshRuns(n_clicks):
    options = [
        {"label": i, "value": i} for i in sorted(os.listdir(data_root)) if "." not in i
    ]
    return options


@app.callback(Output("Amplicon_Library_SC2", "style"), Input("experiment_type", "value"))
def select_primers_sc2(exp_type):
    if exp_type == "SC2-Whole-Genome-Illumina":
        return {"display": "block"}
    else:
        return {"display": "none"}

@app.callback(Output("Amplicon_Library_RSV", "style"), Input("experiment_type", "value"))
def select_primers_rsv(exp_type):
    if exp_type == "RSV-Illumina":
        return {"display": "block"}
    else:
        return {"display": "none"}

def current_version():
    descript_dict = {}
    description_file = f"{dirname(realpath(__file__))}/DESCRIPTION"
    with open(description_file, 'r') as infi:
        for line in infi:
            try:
                descript_dict[line.split(':')[0]]=line.split(":")[1]
            except:
                continue
    return descript_dict['Version'].strip()

@app.callback(
    [Output("select_sample", "options"), Output("select_sample", "value")],
    [
        Input("coverage-heat", "clickData"),
        Input("select_run", "value"),
        Input("irma-results-button", "n_clicks"),
        Input("assembly-button", "n_clicks"),
    ],
)
def select_sample(plotClick, run, n_clicks, a_n_clicks):
    if dash.ctx.triggered_id == "assembly-button":
        return [], ""
    if not run:
        return [], ""  # dash.exceptions.PreventUpdate
    try:
        df = pd.read_json(f"{data_root}/{run}/dash-json/reads.json", orient="split")
    except:
        return [], ""  # None, None
    samples = df["Sample"].unique()
    samples.sort()
    options = [{"label": i, "value": i} for i in samples]
    global previousClick
    if not plotClick:
        value = samples[0]
    elif plotClick == previousClick:
        sample = sample
    elif plotClick["points"][0]["x"] != "all":
        previousClick = plotClick
        value = plotClick["points"][0]["x"]
    return options, value


@app.callback(
    Output("single_sample_figs", "children"),
    [
        Input("select_run", "value"),
        Input("select_sample", "value"),
        Input("irma-results-button", "n_clicks"),
        Input("assembly-button", "n_clicks"),
    ],
)
def single_sample_fig(run, sample, n_clicks, a_n_clicks):
    if dash.ctx.triggered_id == "assembly-button":
        return [html.Div()]  # blank_fig()
    if not run or not sample:
        return html.Div()
    # if not cov_linear_y:
    y_axis_type = "linear"
    # else:
    #    y_axis_type = "log"
    try:
        sankeyfig = pio.read_json(f"{data_root}/{run}/dash-json/readsfig_{sample}.json")
        coveragefig = pio.read_json(
            f"{data_root}/{run}/dash-json/coveragefig_{sample}_{y_axis_type}.json"
        )
    except FileNotFoundError:
        sankeyfig, coveragefig = blank_fig(), blank_fig()
    content = dbc.Row(
        [
            dbc.Col(
                dcc.Graph(
                    figure=sankeyfig,
                    config={
                        "toImageButtonOptions": {
                            "format": "svg",
                            "filename": f"{run}_{sample}_sankey",
                        }
                    },
                ),
                width=4,
                align="start",
            ),
            dbc.Col(
                dcc.Graph(
                    figure=coveragefig,
                    config={
                        "toImageButtonOptions": {
                            "format": "svg",
                            "filename": f"{run}_{sample}_coverage",
                        }
                    },
                ),
                width=8,
                align="center",
            ),
        ]
    )

    return content


@app.callback(
    Output("download_ss", "data"),
    [
        Input("select_run", "value"),
        Input("ss_dl_button", "n_clicks"),
        Input("experiment_type", "value"),
    ],
    prevent_initial_call=True,
)
def download_ss(run, n_clicks, experiment_type):
    global selected_experiment_type
    selected_experiment_type = experiment_type
    if not n_clicks:
        raise dash.exceptions.PreventUpdate
    global dl_ss_clicks
    if n_clicks > dl_ss_clicks:
        dl_ss_clicks = n_clicks
        try:
            generate_samplesheet_xl(run)
            return dcc.send_file(f"{data_root}/{run}/{run}_samplesheet.xlsx")
        except Exception as e:
            print(
                f"----------------------------\nERROR DOWNLOADING SS\n{e}\n--------END OF ERROR--------"
            )
            # get a warning displayed somehow
            # return html.Div(['There was an error processing this run and datatype'])


# from https://dash.plotly.com/dash-core-components/upload


def parse_contents(contents, filename, date, run):
    content_type, content_string = contents.split(",")
    # global ss_df
    decoded = base64.b64decode(content_string)
    try:
        if "csv" in filename:
            # Assume that the user uploaded a CSV file
            ss_df = pd.read_csv(io.StringIO(decoded.decode("utf-8")))
        elif "xls" in filename:
            # Assume that the user uploaded an excel file
            ss_df = pd.read_excel(io.BytesIO(decoded), engine="openpyxl")
            ss_df = ss_df.iloc[:, 0:3]
    except Exception as e:
        print(f"ERROR PARSING SS\n{e}\n--------END OF ERROR--------")
        return html.Div(["There was an error processing this file."])
    for c in ss_df.columns:
        ss_df[c] = ss_df[c].apply(lambda x: str(x))
        # ss_df[c] = ss_df[c].apply(lambda x: x.strip("^M").strip())
    if True in list(ss_df.duplicated("Sample ID")):
        ss_df["duplicated"] = ss_df.duplicated("Sample ID")
        stmnt = f"No duplicate sample IDs allowed. Please edit. Duplicates = {list(ss_df.loc[ss_df['duplicated']==True]['Sample ID'])}"
    elif True in list(ss_df["Sample ID"].str.contains(r"\s")):
        ss_df["spaces"] = ss_df["Sample ID"].str.contains(r"\s")
        stmnt = f"No spaces allowed in Sample IDs. Please edit. Offenders = {list(ss_df.loc[ss_df['spaces']==True]['Sample ID'])}"
    elif True in list(ss_df["Sample ID"].str.contains(r"[\\/]")):
        ss_df["slashes"] = ss_df["Sample ID"].str.contains(r"[\\/]")
        stmnt = f"No '/' or '\\' (slashes) allowed in Sample IDs. Please edit. Offenders = {list(ss_df.loc[ss_df['slashes']==True]['Sample ID'])}"
    elif selected_experiment_type is None:
        stmnt = f'You have not selected what kind of data this run is! Please select from the "What kind of data is this?" dropdown above'
    elif (
        "illumina" in selected_experiment_type.lower() and "Barcode #" in ss_df.columns
    ):
        stmnt = f"You have selected a Nanopore Samplesheet for an Illumina Run!! Please reload with an Illumina Samplesheet"
    elif "ont" in selected_experiment_type.lower() and "Barcode #" not in ss_df.columns:
        stmnt = f"You have selected an Illumina Samplesheet for a Nanopore run!! Please reload with a Nanopore Samplsheet"
    else:
        print(f"saving df")
        ss_df.to_csv(f"{data_root}/{run}/samplesheet.csv", index=False)
        stmnt = ""
    return (
        html.Div(
            [
                html.H5(filename),
                html.H6(datetime.datetime.fromtimestamp(date)),
                dash_table.DataTable(
                    ss_df.to_dict("records"),
                    [{"name": i, "id": i} for i in ss_df.columns],
                ),
                html.Hr(),  # horizontal line
            ]
        ),
        stmnt,
    )


@app.callback(
    [
        Output("samplesheet_errors", "children"),
        Output("upload-data", "contents"),
        Output("upload-data", "filename"),
        Output("upload-data", "last_modified"),
    ],
    Input("upload-data", "contents"),
    Input("select_run", "value"),
    State("upload-data", "filename"),
    State("upload-data", "last_modified"),
)
def update_output(list_of_contents, run, list_of_names, list_of_dates):
    if list_of_contents is not None:
        children = [
            parse_contents(c, n, d, run)
            for c, n, d, run in zip(
                list_of_contents, list_of_names, list_of_dates, itertools.repeat(run)
            )
        ]
        stmnt = children.pop(-1)
        stmnt = [stmnt[-1]]
        return stmnt, None, None, None
    else:
        return html.Div(), None, None, None


def generate_samplesheet_xl(run):
    wb = xl.Workbook()
    ws = wb.active
    # Store possible drop down options
    for r, t in enumerate(["- Control", "+ Control", "Test"]):
        ws[f"Z{r+1}"].value = t
    sample_types = DV(type="list", formula1=f"=Z$1:Z$3")
    ws.add_data_validation(sample_types)
    bar_nums = [int(i[-2:]) for i in glob(f"{data_root}/{run}/fastq_pass/b*")]
    if len(bar_nums) == 0:
        fqs = [
            i.split("/")[-1]
            for i in glob(f"{data_root}/{run}/**/*.fastq*", recursive=True)
        ]
        try:
            ill_samples = list(
                set([re.findall(r".+(?=_S[0-9]+_R[12])", i)[0] for i in fqs])
            )
            if len(ill_samples) < len(fqs) / 2:
                ill_samples = list(
                    set([re.findall(r".+(?=[_/.]R[12])", i)[0] for i in fqs])
                )
        except IndexError:
            ill_samples = list(set([re.findall(r".+(?=[_/.]R[12])", i)[0] for i in fqs]))
        ill_samples.sort()
        print(fqs, ill_samples)
        ws["A1"].value, ws["B1"].value = "Sample ID", "Sample Type"
        row = 2
        for r, s in enumerate(ill_samples):
            ws[f"A{row}"].value = s
            ws[f"B{row}"].value = "Test"
            row += 1
        sample_types.add(f"B2:B{len(ill_samples)+1}")
    else:
        ws["A1"].value, ws["B1"].value, ws["C1"].value = (
            "Barcode #",
            "Sample ID",
            "Sample Type",
        )
        bar_nums.sort()
        barcodes = [f"barcode{i:02}" for i in bar_nums]
        row = 2
        for r, b in enumerate(barcodes):
            ws[f"A{row}"].value = b
            ws[f"C{row}"].value = "Test"
            row += 1
        sample_types.add(f"C2:C{len(barcodes)+1}")
    wb.save(f"{data_root}/{run}/{run}_samplesheet.xlsx")
    print(ws)


@app.callback(
    Output("samplesheet", "children"),
    [
        Input("select_run", "value"),
        Input("upload-data", "contents"),
        Input("experiment_type", "value"),
    ],
)
# only generate if csv already exitsts--otherwise, excel upload/parse functions display ss
def generate_samplesheet(run, upload_data, experiment_type):
    if None in (run, experiment_type):
        return html.Div()
    ss_glob = [
        i
        for i in glob(f"{data_root}/{run}/*.csv*")
        if "samplesheet" in i.lower() or "data" in i.lower()
    ]
    if len(ss_glob) == 1:
        ss_filename = ss_glob[0]
        with open(ss_filename, "rb") as d:
            originalByteList = d.readlines()
        noCarriageReturn = "".join(
            [i.decode().replace("\r", "") for i in originalByteList]
        )
        with open(ss_filename, "w") as d:
            d.write(noCarriageReturn)
        data = pd.read_csv(ss_filename, skip_blank_lines=True).to_dict("records")
        if len(data[0].keys()) > 2:
            if "Illumina" in experiment_type:
                data_columns = [
                    {"id": "Sample ID", "name": "Sample ID"},
                    {
                        "id": "Sample Type",
                        "name": "Sample Type",
                        "presentation": "dropdown",
                    },
                ]
            else:
                data_columns = [
                    {"id": "Barcode #", "name": "Barcode #"},
                    {"id": "Sample ID", "name": "Sample ID"},
                    {
                        "id": "Sample Type",
                        "name": "Sample Type",
                        "presentation": "dropdown",
                    },
                ]
        else:
            data_columns = [
                {"id": "Sample ID", "name": "Sample ID"},
                {"id": "Sample Type", "name": "Sample Type"},
            ]
        ss = dash_table.DataTable(
            id="samplesheet_table",
            columns=data_columns,
            data=data,
            # editable=True,
            merge_duplicate_headers=True,
            export_format="xlsx",
            export_headers="display",
        )
    else:
        return None
    table = ss
    return table


@app.callback(
    Output("assembly-button", "disabled"),
    [
        Input("assembly-button", "n_clicks"),
        Input("select_run", "value"),
        Input("experiment_type", "value"),
        Input("Amplicon_Library", "value"),
        Input("unlock-assembly-button", "n_clicks"),
    ],
)
def run_snake_script_onClick(
    assembly_n_clicks, run, experiment_type, Amplicon_Library, unlock_n_clicks
):
    if dash.ctx.triggered_id == "unlock-assembly-button":
        return False
    ss_glob = [
        i
        for i in glob(f"{data_root}/{run}/samplesheet.csv*")
        if "samplesheet" in i.lower() or "data" in i.lower()
    ]
    if len(ss_glob) != 1:
        return True
    if assembly_n_clicks == None:
        return True
    if dash.ctx.triggered_id == "experiment_type":
        return False
    if dash.ctx.triggered_id == "select_run":
        return True
    if dash.ctx.triggered_id == "assembly-button":
        docker_cmd = "docker exec -w /data spyne bash snake-kickoff "
        docker_cmd += f"{run}/samplesheet.csv "
        docker_cmd += f"{run} "
        docker_cmd += f"{experiment_type} "
        if "sc2-whole-genome-illumina" in experiment_type.lower() or "rsv-illumina" in experiment_type.lower():
            docker_cmd += f"{Amplicon_Library} "
        docker_cmd += f"CLEANUP-FOOTPRINT"
        print(f'launching docker_cmd == "{docker_cmd}"\n\n')
        subprocess.Popen(docker_cmd.split(), close_fds=True)
        if len(glob(f"{data_root}/{run}/{run}_samplesheet.xlsx")) > 0:
            os.remove(
                f"{data_root}/{run}/{run}_samplesheet.xlsx"
            )  # This is only the empty template generated. The used file is saved as samplesheet.csv
        return True


# @app.callback(Output("assembly-button", "disabled"),
#              Input("unlock-assembly-button", "n_clicks"))
# def unlock_assembly_button(n_clicks):
#    return False


@app.callback(
    Output("new-version", "children"), Input("new-version-interval", "n_intervals")
)
def new_version_modal(n_interval):
    if DEPLOY:
        with open("/MIRA/DESCRIPTION", "r") as d:
            current = "".join(d.readlines())
    else:
        with open("DESCRIPTION", "r") as d:
            current = "".join(d.readlines())
    available = requests.get(AVAILABLE_VERSION)
    current = re.findall(r"Version.+(?=\n)", current)[0]
    available = re.findall(r"Version.+(?=\n)", available.text)[0]
    if current >= available:
        return html.Div()
    else:
        modal = dbc.Modal(
            [
                dbc.ModalHeader(
                    html.A(
                        f"A new version of MIRA is available! ",
                        style={"color": "indigo"},
                    )
                ),
                dbc.ModalBody(
                    dcc.Link(
                        "CLICK HERE",
                        href="https://cdcgov.github.io/MIRA/articles/upgrading-mira.html",
                        target="_blank",
                    )
                ),
                dbc.ModalBody(f"Current {current}"),
                dbc.ModalBody(f"Available {available}"),
            ],
            size="lg",
            is_open=True,
            centered=True,
            style={"font-size": "300%", "font-family": "arial"},
        )
        return modal


@app.callback(
    Output("irma-progress", "children"),
    [
        Input("select_run", "value"),
        Input("watch-irma-progress", "value"),
        Input("irma-progress-interval", "n_intervals"),
        Input("assembly-button", "n_clicks"),
    ],
)
def display_irma_progress(run, toggle, n_intervals, n_clicks):
    if not toggle:
        return html.Div()
    if len(glob(f"{data_root}/{run}/*amended_consensus.fasta")) >= 1:
        return html.Div('IRMA is finished! Click "DISPLAY IRMA RESULTS"')
    if (len(glob(f"{data_root}/{run}/dash-json")) == 1) or (
        len(glob(f"{data_root}/{run}/DAIS_ribosome_input.fasta")) == 1
    ):
        if not len(glob(f"{data_root}/{run}/spyne_logs.tar.gz")) == 1:
            return html.Div("Annotating genomes and creating images, please wait...")
        elif len(glob(f"{data_root}/{run}/*amended_consensus.fasta")) == 0:
            return html.Div(
            "Run has failed during annotation and figure creation. If you need further help, please contact us at IDSeqsupport@cdc.gov"
        )
    logs = glob(f"{data_root}/{run}/logs/*irma*out.log")
    if os.path.exists(f"{data_root}/{run}/.snakemake") and len(logs) == 0:
        return html.Div("Data processing has started, please wait...")
    if (len(glob(f"{data_root}/{run}/spyne_logs.tar.gz")) == 1) and (
        len(glob(f"{data_root}/{run}/dash-json")) == 0
    ):
        return html.Div(
            "Run has failed. If you need further help, please contact us at IDSeqsupport@cdc.gov"
        )
    if len(logs) == 0:
        return html.Div("No IRMA data is available")
    log_dic = {}
    for l in logs:
        sample = l.split("/")[-1].split(".")[0]
        with open(l, "r") as d:
            log_dic[sample] = "".join(d.readlines())
    finished_samples = [i for i in log_dic.keys() if "finished!" in log_dic[i].lower()]
    running_samples = {
        i: f"  ".join(j.split("\t"))
        for i, j in log_dic.items()
        if i not in finished_samples
    }
    if (
        len(glob(f"{data_root}/{run}/IRMA/all.fin")) == 1
        or len(glob(f"{data_root}/{run}/spyne_logs.tar.gz")) == 1
    ):
        return html.Div(
            "IRMA has finished running. Once the MIRA tab header has stopped displaying 'Update', click 'Display IRMA Results'"
        )
    df = pd.DataFrame.from_dict(running_samples, orient="index")
    df = df.reset_index()
    try:
        df.columns = ["Sample", "IRMA Stage"]
        out = html.Div(
            [
                html.Div(f"IRMA finished samples: {', '.join(finished_samples)}"),
                html.Div(
                    dash_table.DataTable(
                        columns=[{"name": i, "id": i} for i in df.columns],
                        data=df.to_dict("records"),
                        style_cell={
                            "whiteSpace": "pre-line",
                            "height": "auto",
                            "lineHeight": "15px",
                            "text_align": "left",
                        },
                    )
                ),
            ]
        )
    except ValueError:
        return html.Div()
    return out


def flfor(x, digits):
    return float(f"{x:.{digits}f}")


@app.callback(
    [
        Output("minor_alleles_table", "children"),
        [
            Input("select_run", "value"),
            Input("irma-results-button", "n_clicks"),
            Input("assembly-button", "n_clicks"),
        ],
    ],
)
def alleles_table(run, n_clicks, a_n_clicks):
    if dash.ctx.triggered_id == "assembly-button":
        return [html.Div()]  # blank_fig()
    if not run:
        return [html.Div()]  # raise dash.exceptions.PreventUpdate
    try:
        df = pd.read_json(f"{data_root}/{run}/dash-json/alleles.json", orient="split")
    except:
        return [html.Div()]  # raise dash.exceptions.PreventUpdate
    table = [
        html.Div(
            [
                dash_table.DataTable(
                    columns=[{"name": i, "id": i} for i in df.columns],
                    data=df.to_dict("records"),
                    sort_action="native",
                    filter_action="native",
                    page_size=10,
                    export_format="xlsx",
                    export_headers="display",
                    style_cell={"textAlign": "left", "height": "auto"},
                )
            ]
        )
    ]
    return table


@app.callback(
    [
        Output("indels_table", "children"),
        [
            Input("select_run", "value"),
            Input("irma-results-button", "n_clicks"),
            Input("assembly-button", "n_clicks"),
        ],
    ],
)
def indels_table(run, n_clicks, a_n_clicks):
    if dash.ctx.triggered_id == "assembly-button":
        return [html.Div()]  # blank_fig()
    if not run:
        return [html.Div()]  # raise dash.exceptions.PreventUpdate
    try:
        df = pd.read_json(f"{data_root}/{run}/dash-json/indels.json", orient="split")
    except:
        return [html.Div()]  # raise dash.exceptions.PreventUpdate
    table = [
        html.Div(
            [
                dash_table.DataTable(
                    columns=[{"name": i, "id": i} for i in df.columns],
                    data=df.to_dict("records"),
                    sort_action="native",
                    filter_action="native",
                    page_size=10,
                    export_format="xlsx",
                    export_headers="display",
                    style_cell={"textAlign": "left", "height": "auto"},
                )
            ]
        )
    ]
    return table


@app.callback(
    [
        Output("vars_table", "children"),
        [
            Input("select_run", "value"),
            Input("irma-results-button", "n_clicks"),
            Input("assembly-button", "n_clicks"),
        ],
    ],
)
def vars_table(run, n_clicks, a_n_clicks):
    if dash.ctx.triggered_id == "assembly-button":
        return [html.Div()]  # blank_fig()
    if not run:
        return [html.Div()]  # raise dash.exceptions.PreventUpdate
    try:
        df = pd.read_json(f"{data_root}/{run}/dash-json/dais_vars.json", orient="split")
    except:
        return [html.Div()]  # raise dash.exceptions.PreventUpdate
    table = [
        html.Div(
            [
                dash_table.DataTable(
                    columns=[{"name": i, "id": i} for i in df.columns],
                    data=df.to_dict("records"),
                    sort_action="native",
                    filter_action="native",
                    page_size=10,
                    export_format="xlsx",
                    export_headers="display",
                    style_cell={
                        "whiteSpace": "normal",
                        "textAlign": "left",
                        "height": 40,
                    },
                )
            ]
        )
    ]
    return table


@app.callback(
    Output("demux_fig", "figure"),
    [
        Input("select_run", "value"),
        Input("irma-results-button", "n_clicks"),
        Input("assembly-button", "n_clicks"),
    ],
    prevent_initial_call=True,
)
def barcode_pie(run, n_clicks, a_n_clicks):
    if dash.ctx.triggered_id == "assembly-button":
        return blank_fig()
    try:
        fig = pio.read_json(f"{data_root}/{run}/dash-json/barcode_distribution.json")
    except:
        fig = blank_fig()
    return fig


def negative_qc_statement(run):
    with open(f"{data_root}/{run}/dash-json/qc_statement.json", "r") as d:
        qc_statement_dic = json.load(d)
    statement = [html.Br()]
    for q in ["FAILS QC", "passes QC"]:
        for s, p in qc_statement_dic[q].items():
            if q == "FAILS QC":
                statement.extend(
                    [
                        f"You negative sample ",
                        html.Strong(f'"{s}" FAILS QC ', className="display-7"),
                        f"with {p}% reads mapping to reference.",
                        html.Br(),
                    ]
                )
            else:
                statement.extend(
                    [
                        f'You negative sample "{s}" passes QC with {p}% reads mapping to reference.',
                        html.Br(),
                    ]
                )
    statement.extend([html.Br()])
    return html.P(statement)


@app.callback(
    [Output("irma_neg_statment", "children"), Output("irma_summary", "children")],
    [
        Input("select_run", "value"),
        Input("irma-results-button", "n_clicks"),
        Input("assembly-button", "n_clicks"),
    ],
)
def irma_summary(run, n_clicks, a_n_clicks):
    if dash.ctx.triggered_id == "assembly-button":
        return html.Div(), html.Div()
    try:
        qc_statement = negative_qc_statement(run)
        df = pd.read_json(
            f"{data_root}/{run}/dash-json/irma_summary.json", orient="split"
        )
    except Exception as E:
        qc_statement = html.Div()
    try:
        fill_colors = conditional_color_range_perCol.discrete_background_color_bins(
            df, 8
        )
        table = html.Div(
            [
                dash_table.DataTable(
                    columns=[{"name": i, "id": i} for i in df.columns],
                    data=df.to_dict("records"),
                    sort_action="native",
                    style_data_conditional=fill_colors,
                    export_format="xlsx",
                    export_headers="display",
                    merge_duplicate_headers=True,
                    filter_action="native",
                )
            ]
        )
    except UnboundLocalError:
        table = html.Div()
    return qc_statement, table


@app.callback(
    Output("coverage-heat", "figure"),
    [
        Input("select_run", "value"),
        Input("irma-results-button", "n_clicks"),
        Input("assembly-button", "n_clicks"),
    ],
)
def callback_heatmap(run, n_clicks, a_n_clicks):
    if dash.ctx.triggered_id == "assembly-button":
        return blank_fig()
    if not run:
        return blank_fig()
    try:
        return pio.read_json(f"{data_root}/{run}/dash-json/heatmap.json")
    except FileNotFoundError:
        return blank_fig()


@app.callback(
    Output("pass_fail_heat", "figure"),
    [
        Input("select_run", "value"),
        Input("irma-results-button", "n_clicks"),
        Input("assembly-button", "n_clicks"),
    ],
)
def callback_pass_fail_heatmap(run, n_clicks, a_n_clicks):
    if dash.ctx.triggered_id == "assembly-button":
        return blank_fig()
    if not run:
        return blank_fig()
    try:
        return pio.read_json(f"{data_root}/{run}/dash-json/pass_fail_heatmap.json")
    except FileNotFoundError:
        return blank_fig()


dl_passed_fasta_clicks = 0
dl_failed_fasta_clicks = 0
dl_ss_clicks = 0

flu_numbers = {
    "A": {
        "1": "PB2",
        "2": "PB1",
        "3": "PA",
        "4": "HA",
        "5": "NP",
        "6": "NA",
        "7": "MP",
        "8": "NS",
    },
    "B": {
        "1": "PB1",
        "2": "PB2",
        "3": "PA",
        "4": "HA",
        "5": "NP",
        "6": "NA",
        "7": "MP",
        "8": "NS",
    },
}


@app.callback(
    [
        Output("download_passed_nt_fasta", "data"),
        Output("download_passed_aa_fasta", "data"),
    ],
    [Input("select_run", "value"), Input("fasta_passed_dl_button", "n_clicks")],
    prevent_initial_call=True,
)
def download_passed_fastas(run, n_clicks):
    if not n_clicks:
        raise dash.exceptions.PreventUpdate
    global dl_passed_fasta_clicks
    if n_clicks > dl_passed_fasta_clicks:
        dl_passed_fasta_clicks = n_clicks
        content = open(f"{data_root}/{run}/amended_consensus.fasta").read()
        nt_fastas = dict(
            content=content,
            filename=f"{run}_amended_consensus.fasta",
        )
        content = open(f"{data_root}/{run}/amino_acid_consensus.fasta").read()
        aa_fastas = dict(
            content=content,
            filename=f"{run}_amino_acid_consensus.fasta",
        )
        return nt_fastas, aa_fastas


@app.callback(
    [
        Output("download_failed_nt_fasta", "data"),
        Output("download_failed_aa_fasta", "data"),
    ],
    [Input("select_run", "value"), Input("fasta_failed_dl_button", "n_clicks")],
    prevent_initial_call=True,
)
def download_failed_fastas(run, n_clicks):
    if not n_clicks:
        raise dash.exceptions.PreventUpdate
    global dl_failed_fasta_clicks
    if n_clicks > dl_failed_fasta_clicks:
        dl_failed_fasta_clicks = n_clicks
        content = open(f"{data_root}/{run}/failed_amended_consensus.fasta").read()
        nt_fastas = dict(
            content=content,
            filename=f"{run}_amended_consensus_FAILED.fasta",
        )
        content = open(f"{data_root}/{run}/failed_amino_acid_consensus.fasta").read()
        aa_fastas = dict(
            content=content,
            filename=f"{run}_amino_acid_consensus_FAILED.fasta",
        )
        return nt_fastas, aa_fastas


########################################################
###################### LAYOUT ##########################
########################################################
SIDEBAR_STYLE = {
    "position": "fixed",
    "top": 0,
    "left": 0,
    "bottom": 0,
    "width": "22rem",
    "padding": "2rem 1rem",
    "background-color": "#f8f9fa",
}

# the styles for the main content position it to the right of the sidebar and
# add some padding.
CONTENT_STYLE = {
    "margin-left": "24rem",
    "margin-right": "2rem",
    "padding": "2rem 1rem",
}

sidebar = html.Div(
    [
        html.A(
            href="https://cdcgov.github.io/MIRA/articles/running-mira.html",
            target="_blank",
            children=[
                html.Img(
                    src=app.get_asset_url("mira-logo-midjourney_20230526_rmbkgnd.png"),
                    height=125,
                    width=125,
                )
            ],
        ),
        html.H2(f"MIRA v{current_version()}", className="display-4"),
        html.P(
            [
                "Influenza, SARS-CoV-2, and RSV sequence assembly with ",
                dcc.Link(
                    "IRMA",
                    href="https://bmcgenomics.biomedcentral.com/articles/10.1186/s12864-016-3030-6",
                    target="_blank",
                ),
                " the Iterative Refinement Meta Assembler",
            ],
            className="display-8",
        ),
        html.Hr(),
        dbc.Nav(
            [
                dbc.NavLink(
                    "Samplesheet", href="#samplesheet_head", external_link=True
                ),
                dbc.NavLink(
                    "Barcode Assignment", href="#demux_head", external_link=True
                ),
                dbc.NavLink("Automatic QC", href="#auto_qc_head", external_link=True),
                dbc.NavLink("IRMA Summary", href="#irma_head", external_link=True),
                dbc.NavLink(
                    "Reference Coverage", href="#coverage_head", external_link=True
                ),
                dbc.NavLink(
                    "Reference Variants", href="#variants_head", external_link=True
                ),
                dbc.NavLink("Minor SNVs", href="#alleles_head", external_link=True),
                dbc.NavLink("Minor Indels", href="#indels_head", external_link=True),
                dbc.NavLink("Download Fastas", href="#dl_fastas", external_link=True),
            ],
            vertical=True,
            pills=True,
        ),
    ],
    style=SIDEBAR_STYLE,
)


def blank_fig():
    fig = go.Figure(go.Scatter(x=[], y=[]))
    fig.update_layout(template=None)
    fig.update_xaxes(showgrid=False, showticklabels=False, zeroline=False)
    fig.update_yaxes(showgrid=False, showticklabels=False, zeroline=False)
    return fig


# import dash_uploader as du
content = html.Div(
    id="page-content",
    style=CONTENT_STYLE,
    children=[html.Div(id="new-version")]
    + [
        dcc.Interval(id="new-version-interval", interval=check_version_interval)
    ]  # interval in milliseconds
    + [
        dbc.Row(
            [
                dbc.Col(
                    dcc.Dropdown(
                        id="select_run",
                        options=[
                            {"label": i, "value": i}
                            for i in sorted(os.listdir(data_root))
                            if "." not in i
                        ],
                        placeholder="Select sequencing run: ",
                    ),
                    width=4,
                ),
                dbc.Col(
                    dbc.Button(
                        "Refresh Run Listing", id="run-refresh-button", n_clicks=0
                    )
                ),
            ]
        )
    ]
    + [
        dbc.Row(
            dcc.Dropdown(
                [
                    {"label": "Flu-ONT", "value": "Flu-ONT"},
                    {"label": "SC2-Spike-Only-ONT", "value": "SC2-Spike-Only-ONT"},
                    {"label": "Flu-Illumina", "value": "Flu-Illumina"},
                    {"label": "SC2-Whole-Genome-ONT", "value":"SC2-Whole-Genome-ONT"},
                    {"label": "SC2-Whole-Genome-Illumina", "value": "SC2-Whole-Genome-Illumina"},
                    {"label": "RSV-Illumina", "value": "RSV-Illumina"},
                    {"label": "RSV-ONT", "value": "RSV-ONT",},
                ],
                id="experiment_type",
                placeholder="What kind of data is this?",
            )
        )
    ]
    + [  # html.Div(id="Amplicon_Library")
        dbc.Row(
            dcc.Dropdown(
                [
                    {"label": "Artic V3", "value": "articv3"},
                    {"label": "Artic V4", "value": "articv4"},
                    {"label": "Artic V4.1", "value": "articv4.1"},
                    {"label": "Artic V5.3.2", "value": "articv5.3.2"},
                    {"label": "Qiagen QIAseq", "value": "qiagen"},
                    {"label": "xGen™ SARS-CoV-2 Amplicon Panel", "value": "swift"},
                    {
                        "label": "xGen™ SARS-CoV-2 Amplicon Panel (CDC customized)",
                        "value": "swift_211206",
                    },
                    {"label": "VarSkip", "value": "varskip"},
                    {"label": "None", "value": ""},
                ],  # add handling here for no primers used
                id="Amplicon_Library_SC2",
                placeholder="For Illumina SC2, which primer schema was used?",
            )
        )
    ]
    + [
        dbc.Row(
            dcc.Dropdown(
                [
                    {"label": "RSV CDC 8 amplicon 230901", "value": "RSV_CDC_8amplicon_230901"},
                ],  # add handling here for no primers used
                id="Amplicon_Library_RSV",
                placeholder="For Illumina RSV, which primer schema was used?",
            )
        )
    ]
    + [html.P("Samplesheet", id="samplesheet_head", className="display-6")]
    + [
        html.Div(
            [
                dbc.Button("Download Samplesheet Template", id="ss_dl_button"),
                dcc.Download(id="download_ss"),
            ]
        )
    ]
    + [
        html.Div(
            [
                dcc.Upload(
                    id="upload-data",
                    children=html.Div(
                        [
                            "Drag and Drop your ",
                            html.B("Samplesheet"),
                            " or Click and Select the File",
                        ]
                    ),
                    style={
                        "width": "100%",
                        "height": "60px",
                        "lineHeight": "60px",
                        "borderWidth": "1px",
                        "borderStyle": "dashed",
                        "borderRadius": "5px",
                        "textAlign": "center",
                        "margin": "10px",
                    },
                    # Allow multiple files to be uploaded
                    multiple=True,
                ),
                html.Div(id="output-data-upload"),
            ]
        )
    ]
    + [html.Br()]
    + [html.Div(id="samplesheet")]
    + [html.Div(id="samplesheet_errors")]
    + [html.Br()]
    + [
        dbc.Row(
            [
                dbc.Col(
                    dbc.Button(
                        "Start Genome Assembly",
                        id="assembly-button",
                        n_clicks=0,
                        disabled=True,
                    ),
                    width="auto",
                ),
                dbc.Col(
                    dbc.Button(
                        "Unlock Assembly Button",
                        id="unlock-assembly-button",
                        outline=True,
                        color="secondary",
                        className="me-1",
                    ),
                    width="auto",
                ),
            ]
        ),
        html.Div(id="output-container-button"),
        dcc.Interval(id="irma-progress-interval", interval=3000),
        html.Div(
            [
                dbc.Row(
                    dbc.Col(
                        daq.ToggleSwitch(
                            id="watch-irma-progress",
                            label="Watch IRMA progress",
                            labelPosition="right",
                            value=False,
                        ),
                        width=2,
                    )
                )
            ]
        ),
        html.Div(id="irma-progress"),
    ]
    + [
        html.Div(
            dbc.Button("Display IRMA results", id="irma-results-button", n_clicks=0),
        )
    ]
    + [html.P("Barcode Assignment", id="demux_head", className="display-6")]
    + [
        html.Div(
            [
                dbc.Row(
                    dbc.Col(
                        dcc.Loading(
                            id="demux-loading",
                            type="cube",
                            children=[
                                dcc.Graph(
                                    id="demux_fig",
                                    figure=blank_fig(),
                                    config={
                                        "toImageButtonOptions": {
                                            "format": "svg",
                                            "filename": "demux",
                                        }
                                    },
                                )
                            ],
                        ),
                        width={"size": 6, "offset": 3},
                    )
                ),
            ]
        )
    ]
    + [html.Br()]
    + [
        html.P(
            "Automatic Quality Control Decisions",
            id="auto_qc_head",
            className="display-6",
        )
    ]
    + [
        html.Div(
            [
                dbc.Row(
                    [
                        html.Div(id="irma_neg_statment"),
                        dbc.Col(
                            dcc.Loading(
                                id="qcheat-loading",
                                type="cube",
                                children=[
                                    dcc.Graph(
                                        id="pass_fail_heat",
                                        figure=blank_fig(),
                                        config={
                                            "toImageButtonOptions": {
                                                "format": "svg",
                                                "filename": "passfail_heatmap",
                                            }
                                        },
                                    ),
                                ],
                            ),
                            width=11,
                            align="end",
                        ),
                    ]
                )
            ]
        )
    ]
    + [html.P("IRMA Summary", id="irma_head", className="display-6")]
    + [html.Br()]
    + [dbc.Row([html.Div(id="irma_summary")])]
    + [html.Br()]
    + [html.P("Reference Coverage", id="coverage_head", className="display-6")]
    + [html.Br()]
    + [
        html.Div(
            [
                dbc.Row(
                    [
                        dbc.Col(
                            dcc.Loading(
                                id="coverageheat-loading",
                                type="cube",
                                children=[
                                    dcc.Graph(
                                        id="coverage-heat",
                                        figure=blank_fig(),
                                        config={
                                            "toImageButtonOptions": {
                                                "format": "svg",
                                                "filename": "coverage_heatmap",
                                            }
                                        },
                                    )
                                ],
                            ),
                            width=11,
                            align="end",
                        )
                    ]
                ),
                dcc.Dropdown(id="select_sample"),
                html.Div(id="single_sample_figs"),
            ]
        )
    ]
    + [html.Br()]
    + [html.P("Reference Variants", id="variants_head", className="display-6")]
    + [html.Br()]
    + [html.P( ["Non-amino-acid variants ", dcc.Link(
                    "key",
                    href="https://cdcgov.github.io/MIRA/articles/running-mira.html#special-translated-characters",
                    target="_blank",
                ), ], className="display-8")
                ]
    + [html.Br()]
    + [dbc.Row(html.Div(id="vars_table"))]
    + [html.Br()]
    + [html.P("Minor SNVs", id="alleles_head", className="display-6")]
    + [html.Br()]
    + [dbc.Row(html.Div(id="minor_alleles_table"))]
    + [html.Br()]
    + [
        html.P(
            "Minor Insertions and Deletions", id="indels_head", className="display-6"
        )
    ]
    + [html.Br()]
    + [dbc.Row(html.Div(id="indels_table"))]
    + [html.Br()]
    + [html.P(id="dl_fastas")]
    + [
        html.Div(
            [
                dbc.Button(
                    "Download Fastas", id="fasta_passed_dl_button", color="primary"
                ),
                dcc.Download(id="download_passed_nt_fasta"),
                dcc.Download(id="download_passed_aa_fasta"),
                dbc.Button(
                    "Download Failed Fastas",
                    id="fasta_failed_dl_button",
                    color="secondary",
                ),
                dcc.Download(id="download_failed_nt_fasta"),
                dcc.Download(id="download_failed_aa_fasta"),
            ]
        )
    ],
)

app.layout = html.Div([sidebar, content])
app.update_title = None
####################################################
####################### MAIN #######################
####################################################
if __name__ == "__main__":
    app.run_server(host="0.0.0.0", debug=DEBUG)
