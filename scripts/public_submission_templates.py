import pandas as pd
import yaml
import subprocess
import re
from os.path import dirname
import time
import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows

repo_root = dirname(dirname(__file__))
################################################################################
#### Versions
################################################################################
with open(f"{repo_root}/DESCRIPTION", "r") as d:
    mira_version = "".join(d.readlines())
    mira_version = re.findall(r"Version.+(?=\n)", mira_version)[0]
    mira_version = f"v{mira_version.split()[-1]}"

irma_version = subprocess.run(
    "docker exec -it -w /data irma IRMA", shell=True, capture_output=True
)
irma_version = re.findall(r"(?<=v)\d+.\d+.\d+", irma_version.stdout.decode())[0]
irma_version = f"v{irma_version}"

################################################################################
#### Fields
################################################################################
user_required_fields_submitter = {
    "NCBI": [
        "Username",
        "Password",
        "First_name",
        "Last_name",
        "Email",
        "Authors_list",
    ],
    "GISAID": ["Client-Id", "Username", "Password", "Authors_list"],
}

################################################################################
#### Functions
################################################################################
def generate_submitter_template_xlsx(urfs=user_required_fields_submitter):
    df = pd.DataFrame(columns=["Repository", "Field", "Value"])
    for repo, fields in urfs.items():
        for field in fields:
            df = df.append(
                {"Repository": repo, "Field": field, "Value": ""}, ignore_index=True
            )
    df.to_excel("submitter_config_template.xlsx", index=False, engine="openpyxl")
    return df


def generate_samples_template_xlsx(
    passed_df,
    rundir,
    #urfs=user_required_fields_submission,
    #uofs=user_optional_fields_submission,
    #mrfs=mira_required_fields_submission
):
    with open(f"{rundir}/../.seqsender/config.yaml") as y:
        config = yaml.safe_load(y)
    print(f"zip = {zip(passed_df)}")
    passed_df = passed_df.loc[
        (passed_df["Pass/Fail Reason"] == "Pass")
        | (
            (passed_df["Pass/Fail Reason"].str.contains("Premature stop codon"))
            & (~passed_df["Pass/Fail Reason"].str.contains(";", na=False))
        )
    ]
    # Identify if Flu segments are in dataframe as references
    print(f"{passed_df['Reference']}")
    ll = [
        [s for s in ["PB2", "PB1", "PA", "HA", "NP", "NA", "M", "NS"] if s in l]
        for l in passed_df["Reference"]
    ]
    if len([i for l in ll for i in l]) > 0:
        virus = "FLU"
    elif "SARS-CoV-2" in passed_df["Reference"]:
        virus = "SC2"
    else:
        print(f"{[i for l in ll for i in l]} -- {passed_df['Reference']}")
    ####################################################################################################################################################
    ## JUST GISAID (2023-07-24)
    repo = "GISAID"
    if virus == "FLU":
        p2 = passed_df.loc[:, ["Sample", "Reference"]]
        p2["fasta_header"] = p2.apply(lambda x: f"{x.Sample}|{x.Reference}", axis=1)
        p2["Segment"] = p2["Reference"].apply(lambda x: x.split("_")[1])
        p2 = p2.pivot(columns="Segment", values="fasta_header", index="Sample")
        p2["Subtype"] = p2.fillna('').apply(
            lambda x: "".join(
                [re.findall(r"H\d|$|B", x.HA)[0], re.findall(r"N\d|$", x.NA)[0]]
            ),
            axis=1,
        )
        p2 = p2.reset_index()
        p2 = p2.rename(
            columns={
                "HA": "Seq_Id (HA)",
                "MP": "Seq_Id (MP)",
                "NA": "Seq_Id (NA)",
                "NP": "Seq_Id (NP)",
                "NS": "Seq_Id (NS)",
                "PA": "Seq_Id (PA)",
                "PB1": "Seq_Id (PB1)",
                "PB2": "Seq_Id (PB2)",
                "Sample": "Submitting_Sample_Id",
            }
        )
        if repo == "GISAID":
            g_f_wb = xl.reader.excel.load_workbook(
                f"{repo_root}/lib/gisaid_batch_uploader_flu.xlsx"
            )
            g_f_df = pd.DataFrame(columns=next(g_f_wb.active.values))
            g_f_df = pd.concat([g_f_df, p2])
            g_f_df["Note"] = f"Assembled with MIRA_{mira_version}-IRMA_{irma_version}"
            g_f_df["Authors"] = ";".join(
                f"{i.strip().split()[-1]},{i.strip().split()[0]}"
                for i in config["Submission"][repo]["Authors_list"].split(",")
            )
            for r in dataframe_to_rows(g_f_df, index=False, header=False):
                g_f_wb.active.append(r)
            run = rundir.split("/")[-1]
            g_f_wb.save(f"{rundir}/{run}_submission_draft.xlsx")

    #    # ADD REQUIRED AND OPTIONAL FIELDS
    #    p2 = p2.reindex(
    #        columns=list(
    #            p2.columns.union(set(urfs[virus][repo]))
    #            .union(set(mrfs[virus][repo]))
    #            .union(set(uofs[virus][repo]))
    #        )
    #    )
    # p2["note"] = f"Assembled with MIRA_{mira_version}-IRMA_{irma_version}"
    # Ymd = time.strftime("%Y%m%d", time.localtime())
    # p2["submission_title"] = f"{repo}_{virus}_{Ymd}"
    # if rundir:
    #    run = rundir.split("/"[-1])
    #    p2.to_excel(f"{rundir}/{run}_submission_draft.xlsx", engine="openpyxl", index=False)
    # else:
    #    p2.to_excel(f"submission_draft.xlsx", engine="openpyxl", index=False)
    #return p2

def isolate_name(col_headers, subtype, location, province, host, sub_sample_id, col_date):
    ll=[[s for s in ["PB2", "PB1", "PA", "HA", "NP", "NA", "M", "NS"] if s in l] for l
 in col_headers]
    if len([i for l in ll for i in l]) == 10:
        virus = "FLU"
    elif "SARS-CoV-2" in passed_df["Reference"]:
        virus = "SC2"
    if virus == 'FLU':
        #low_loc = lower(location.replace(' ',''))
        #if low_loc == 'usa' or low_loc == 'us' or 'unitedstates' in low_loc:
        if province is None:
            name_loc = location
        else:
            name_loc = province
        name_year = col_date.split('-')[0]
        if subtype == 'B':
            name_type = subtype
        else:
            name_type = 'A'
        norm_host = host.lower().replace(' ','')
        if norm_host == 'human':
            return f"{name_type}/{name_loc}/{sub_sample_id}/{name_year}"
        else:
            return f"{name_type}/{host}/{name_loc}/{sub_sample_id}/{name_year}"

def generate_isolate_names(rundir, draft_xlsx):
    d_wb = xl.reader.excel.load_workbook(draft_xlsx)
    d_df = pd.DataFrame(d_wb.active.values)
    d_df.columns = d_df.iloc[0]
    d_df = d_df.loc[1:]
    d_df_cols = list(d_df.columns)
    d_df['Isolate_Name'] = d_df.apply(lambda x: isolate_name(d_df_cols, x.Subtype, x.Location, x.province, x.Host, x.Submitting_Sample_Id, x.Collection_Date),axis=1)
    run = rundir.split("/")[-1]
    for r,r_d in enumerate(dataframe_to_rows(d_df, index=False, header=False)):
        for c, c_d in enumerate(r_d):
            d_wb.active.cell(r+2,c+1).value = c_d 
    d_wb.save(f"{rundir}/{run}_submission_final.xlsx")
        



#
#user_required_fields_submission = {
#    "FLU": {
#        "NCBI": [
#            "submission_title",
#            "collection_date",
#            "spuid",
#            "spuid_namespace",
#            "description_title",
#            "organism_name",
#            "bioproject",
#            "bs-collected_by",
#            "bs-collection_date",
#            "bs-geo_loc_name",
#            "bs-host",
#            "bs-host_disease",
#            "bs-strain",
#            "bs-isolation_source",
#            "bs-lat_lon",
#            "sra-file_location",
#            "sra-file_path",
#            "sra-data_type",
#            "sra-instrument_model",
#            "sequence_ID",
#            "src-collection_date",
#            "src-country",
#            "src-host",
#            "src-isolation_source",
#        ],
#        "GISAID": [
#            "submission_title",
#            "collection_date",
#            "isolate_name",
#            "location",
#            "host",
#            "originating_lab_id",
#        ],
#    },
#    "SC2": {
#        "NCBI": [
#            "submission_title",
#            "collection_date",
#            "spuid",
#            "spuid_namespace",
#            "description_title",
#            "organism_name",
#            "bioproject",
#            "bs-collected_by",
#            "bs-collection_date",
#            "bs-geo_loc_name",
#            "bs-host",
#            "bs-host_disease",
#            "bs-isolate",
#            "bs-isolation_source",
#            "sra-file_location",
#            "sra-file_path",
#            "sra-data_type",
#            "sra-instrument_model",
#            "sequence_ID",
#            "src-collection_date",
#            "src-country",
#            "src-host",
#            "src-isolation_source",
#            "src-isolate",
#        ],
#        "GISAID": [
#            "submission_title",
#            "collection_date",
#            "originating_lab_id",
#            "covv_collection_date",
#            "covv_gender",
#            "covv_host",
#            "covv_location",
#            "covv_orig_lab",
#            "covv_orig_lab_address",
#            "covv_treatment",
#            "covv_passage",
#            "covv_patient_age",
#            "covv_patient_status",
#            "covv_virus_name",
#        ],
#    },
#}
#
#user_optional_fields_submission = {
#    "FLU": {
#        "NCBI": [
#            "bs-host_sex",
#            "bs-host_age",
#            "sra-platform",
#            "sra-loader",
#            "src-passage",
#        ],
#        "GISAID": [
#            "lineage",
#            "passage_history",
#            "province",
#            "subprovince",
#            "location_additional_info",
#            "submitting_sample_id",
#            "originating_sample_id",
#            "collection_month",
#            "antigen_char",
#            "Adamantanes_Resistance_geno",
#            "Oseltamivir_Resistance_geno",
#            "Zanamivir_Resistance_geno",
#            "Peramivir_Resistance_geno",
#            "Other_Resistance_geno",
#            "Adamantanes_Resistance_pheno",
#            "Oseltamivir_Resistance_pheno",
#            "Zanamivir_Resistance_pheno",
#            "Peramivir_Resistance_pheno",
#            "Other_Resistance_pheno",
#            "Host_Age",
#            "Host_Age_Unit",
#            "Host_Gender",
#            "Health_Status",
#        ],
#    },
#    "SC2": {
#        "NCBI": [
#            "bs-host_sex",
#            "bs-host_age",
#            "sra-platform",
#            "sra-loader",
#            "src-passage",
#        ],
#        "GISAID": [],
#    },
#}
#
#mira_required_fields_submission = {
#    "FLU": {
#        "NCBI": [
#            "organism",
#            "sra-library_name",
#            "sra-library_strategy",
#            "sra-library_source",
#            "sra-library_selection",
#            "sra-library_layout",
#            "sra-library_construction_protocol",
#            "src-strain",
#            "src-serotype",
#            "cmt-StructuredCommentPrefix",
#            "cmt-Assembly-Method",
#            "cmt-Coverage",
#            "cmt-Sequencing-Technology",
#            "cmt-StructuredCommentSuffix",
#        ],
#        "GISAID": [
#            "Subtype",
#            "Seq_Id (HA)",
#            "Seq_Id (HE)",
#            "Seq_Id (MP)",
#            "Seq_Id (NA)",
#            "Seq_Id (NP)",
#            "Seq_Id (NS)",
#            "Seq_Id (P3)",
#            "Seq_Id (PA)",
#            "Seq_Id (PB1)",
#            "Seq_Id (PB2)",
#            "note",
#        ],
#    },
#    "SC2": {
#        "NCBI": [
#            "organism",
#            "sra-library_name",
#            "sra-library_strategy",
#            "sra-library_source",
#            "sra-library_selection",
#            "sra-library_layout",
#            "sra-library_construction_protocol",
#            "src-organism",
#        ],
#        "GISAID": ["covv_coverage", "covv_seq_technology", "covv_virus_name"],
#    },
#}

